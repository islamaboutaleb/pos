# -*- coding: utf-8 -*-
"""
نظام إدارة الصيانة الوقائية للكاميرات والمهام العامة
Flask + SQLite
"""
import os
import io
import sqlite3
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, request, jsonify, session, send_file, render_template, g
)
from werkzeug.security import check_password_hash, generate_password_hash

from seed_data import get_connection, init_db, DB_PATH

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "camera-maintenance-secret-key-change-me")
app.config["JSON_AS_ASCII"] = False

STATUS_LIST = ["قيد الانتظار", "قيد التنفيذ", "مكتمل"]
MAINTENANCE_KINDS = ["وقائية", "تنظيف", "طارئة"]
TASK_TYPES = ["normal", "camera"]


# --------------------------------------------------------------------------
# قاعدة البيانات: اتصال لكل طلب
# --------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = get_connection()
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# --------------------------------------------------------------------------
# حماية المسارات
# --------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "يجب تسجيل الدخول أولاً"}), 401
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "يجب تسجيل الدخول أولاً"}), 401
        if session.get("role") != "admin":
            return jsonify({"error": "هذا الإجراء مخصص للمدير فقط"}), 403
        return f(*args, **kwargs)
    return wrapper


def gen_task_code(db, task_type):
    prefix = "CAM" if task_type == "camera" else "TSK"
    today = date.today().strftime("%y%m%d")
    cur = db.execute(
        "SELECT COUNT(*) AS c FROM tasks WHERE task_code LIKE ?",
        (f"{prefix}-{today}-%",),
    )
    seq = cur.fetchone()["c"] + 1
    return f"{prefix}-{today}-{seq:03d}"


# --------------------------------------------------------------------------
# صفحات الواجهة
# --------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# --------------------------------------------------------------------------
# المصادقة
# --------------------------------------------------------------------------
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)
    ).fetchone()

    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "اسم المستخدم أو كلمة المرور غير صحيحة"}), 401

    session["user_id"] = user["id"]
    session["role"] = user["role"]
    session["full_name"] = user["full_name"]
    session["technician_id"] = user["technician_id"]

    return jsonify({
        "id": user["id"],
        "username": user["username"],
        "full_name": user["full_name"],
        "role": user["role"],
        "technician_id": user["technician_id"],
    })


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me", methods=["GET"])
def me():
    if "user_id" not in session:
        return jsonify({"authenticated": False})
    return jsonify({
        "authenticated": True,
        "id": session["user_id"],
        "full_name": session["full_name"],
        "role": session["role"],
        "technician_id": session.get("technician_id"),
    })


# --------------------------------------------------------------------------
# المناطق والفروع
# --------------------------------------------------------------------------
@app.route("/api/zones", methods=["GET"])
@login_required
def get_zones():
    db = get_db()
    zones = db.execute("SELECT * FROM zones ORDER BY sort_order, id").fetchall()
    result = []
    for z in zones:
        branches = db.execute(
            "SELECT * FROM branches WHERE zone_id = ? ORDER BY sort_order, id",
            (z["id"],),
        ).fetchall()
        result.append({
            "id": z["id"],
            "name": z["name"],
            "branches": [{"id": b["id"], "name": b["name"]} for b in branches],
        })
    return jsonify(result)


@app.route("/api/zones", methods=["POST"])
@admin_required
def add_zone():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "اسم المنطقة مطلوب"}), 400
    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO zones (name, sort_order) "
            "VALUES (?, (SELECT COALESCE(MAX(sort_order), 0) + 1 FROM zones))",
            (name,),
        )
        db.commit()
        return jsonify({"id": cur.lastrowid, "name": name, "branches": []}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": "هذه المنطقة موجودة مسبقاً"}), 409


@app.route("/api/zones/<int:zone_id>", methods=["PUT"])
@admin_required
def update_zone(zone_id):
    db = get_db()
    zone = db.execute("SELECT * FROM zones WHERE id = ?", (zone_id,)).fetchone()
    if not zone:
        return jsonify({"error": "المنطقة غير موجودة"}), 404
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "اسم المنطقة مطلوب"}), 400
    try:
        db.execute("UPDATE zones SET name = ? WHERE id = ?", (name, zone_id))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "هذا الاسم مستخدم لمنطقة أخرى"}), 409
    return jsonify({"id": zone_id, "name": name})


@app.route("/api/zones/<int:zone_id>", methods=["DELETE"])
@admin_required
def delete_zone(zone_id):
    db = get_db()
    zone = db.execute("SELECT * FROM zones WHERE id = ?", (zone_id,)).fetchone()
    if not zone:
        return jsonify({"error": "المنطقة غير موجودة"}), 404
    task_count = db.execute(
        "SELECT COUNT(*) AS c FROM tasks WHERE zone_id = ?", (zone_id,)
    ).fetchone()["c"]
    if task_count > 0:
        return jsonify({
            "error": f"لا يمكن حذف هذه المنطقة لوجود {task_count} مهمة مرتبطة بها. "
                     f"يجب حذف أو إعادة توجيه هذه المهام أولاً."
        }), 409
    db.execute("DELETE FROM zones WHERE id = ?", (zone_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# الفروع (تضاف ضمن منطقة محددة)
# --------------------------------------------------------------------------
@app.route("/api/zones/<int:zone_id>/branches", methods=["POST"])
@admin_required
def add_branch(zone_id):
    db = get_db()
    zone = db.execute("SELECT * FROM zones WHERE id = ?", (zone_id,)).fetchone()
    if not zone:
        return jsonify({"error": "المنطقة غير موجودة"}), 404
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "اسم الفرع مطلوب"}), 400
    cur = db.execute(
        "INSERT INTO branches (zone_id, name, sort_order) "
        "VALUES (?, ?, (SELECT COALESCE(MAX(sort_order), 0) + 1 FROM branches WHERE zone_id = ?))",
        (zone_id, name, zone_id),
    )
    db.commit()
    return jsonify({"id": cur.lastrowid, "name": name, "zone_id": zone_id}), 201


@app.route("/api/branches/<int:branch_id>", methods=["PUT"])
@admin_required
def update_branch(branch_id):
    db = get_db()
    branch = db.execute("SELECT * FROM branches WHERE id = ?", (branch_id,)).fetchone()
    if not branch:
        return jsonify({"error": "الفرع غير موجود"}), 404
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "اسم الفرع مطلوب"}), 400
    db.execute("UPDATE branches SET name = ? WHERE id = ?", (name, branch_id))
    db.commit()
    return jsonify({"id": branch_id, "name": name})


@app.route("/api/branches/<int:branch_id>", methods=["DELETE"])
@admin_required
def delete_branch(branch_id):
    db = get_db()
    branch = db.execute("SELECT * FROM branches WHERE id = ?", (branch_id,)).fetchone()
    if not branch:
        return jsonify({"error": "الفرع غير موجود"}), 404
    db.execute("DELETE FROM branches WHERE id = ?", (branch_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# الفنيون
# --------------------------------------------------------------------------
@app.route("/api/technicians", methods=["GET"])
@login_required
def get_technicians():
    db = get_db()
    if request.args.get("all") == "1" and session.get("role") == "admin":
        techs = db.execute("SELECT * FROM technicians ORDER BY name").fetchall()
    else:
        techs = db.execute(
            "SELECT * FROM technicians WHERE is_active = 1 ORDER BY name"
        ).fetchall()

    result = []
    for t in techs:
        linked_user = db.execute(
            "SELECT username FROM users WHERE technician_id = ? AND is_active = 1",
            (t["id"],),
        ).fetchone()
        result.append({
            "id": t["id"],
            "name": t["name"],
            "phone": t["phone"],
            "is_active": bool(t["is_active"]),
            "linked_username": linked_user["username"] if linked_user else None,
        })
    return jsonify(result)


@app.route("/api/technicians", methods=["POST"])
@admin_required
def add_technician():
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    if not name:
        return jsonify({"error": "اسم الفني مطلوب"}), 400
    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO technicians (name, phone) VALUES (?, ?)", (name, phone)
        )
        db.commit()
        return jsonify({"id": cur.lastrowid, "name": name, "phone": phone, "is_active": True, "linked_username": None}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": "هذا الفني مسجل مسبقاً"}), 409


@app.route("/api/technicians/<int:tech_id>", methods=["PUT"])
@admin_required
def update_technician(tech_id):
    db = get_db()
    tech = db.execute("SELECT * FROM technicians WHERE id = ?", (tech_id,)).fetchone()
    if not tech:
        return jsonify({"error": "الفني غير موجود"}), 404

    data = request.get_json(force=True) or {}
    fields = {}

    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "اسم الفني مطلوب"}), 400
        fields["name"] = name

    if "phone" in data:
        fields["phone"] = (data.get("phone") or "").strip()

    if "is_active" in data:
        fields["is_active"] = int(bool(data.get("is_active")))

    if fields:
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        try:
            db.execute(
                f"UPDATE technicians SET {set_clause} WHERE id = ?",
                list(fields.values()) + [tech_id],
            )
            db.commit()
        except sqlite3.IntegrityError:
            return jsonify({"error": "هذا الاسم مستخدم لفني آخر"}), 409

    row = db.execute("SELECT * FROM technicians WHERE id = ?", (tech_id,)).fetchone()
    linked_user = db.execute(
        "SELECT username FROM users WHERE technician_id = ? AND is_active = 1", (tech_id,)
    ).fetchone()
    return jsonify({
        "id": row["id"], "name": row["name"], "phone": row["phone"],
        "is_active": bool(row["is_active"]),
        "linked_username": linked_user["username"] if linked_user else None,
    })


@app.route("/api/technicians/<int:tech_id>", methods=["DELETE"])
@admin_required
def delete_technician(tech_id):
    db = get_db()
    tech = db.execute("SELECT * FROM technicians WHERE id = ?", (tech_id,)).fetchone()
    if not tech:
        return jsonify({"error": "الفني غير موجود"}), 404

    linked_user = db.execute(
        "SELECT username FROM users WHERE technician_id = ?", (tech_id,)
    ).fetchone()
    if linked_user:
        return jsonify({
            "error": f"لا يمكن حذف هذا الفني لأنه مرتبط بحساب المستخدم '{linked_user['username']}'. "
                     f"قم بحذف أو تعديل ربط الحساب أولاً."
        }), 409

    db.execute("DELETE FROM technicians WHERE id = ?", (tech_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# المستخدمون (حسابات المديرين والفنيين)
# --------------------------------------------------------------------------
USER_SELECT = """
    SELECT u.*, t.name AS technician_name
    FROM users u
    LEFT JOIN technicians t ON t.id = u.technician_id
"""


def serialize_user(row):
    return {
        "id": row["id"],
        "username": row["username"],
        "full_name": row["full_name"],
        "role": row["role"],
        "technician_id": row["technician_id"],
        "technician_name": row["technician_name"],
        "is_active": bool(row["is_active"]),
        "created_at": row["created_at"],
    }


def count_active_admins(db, exclude_id=None):
    q = "SELECT COUNT(*) AS c FROM users WHERE role = 'admin' AND is_active = 1"
    params = []
    if exclude_id:
        q += " AND id != ?"
        params.append(exclude_id)
    return db.execute(q, params).fetchone()["c"]


@app.route("/api/users", methods=["GET"])
@admin_required
def list_users():
    db = get_db()
    rows = db.execute(USER_SELECT + " ORDER BY u.created_at DESC").fetchall()
    return jsonify([serialize_user(r) for r in rows])


@app.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    data = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    full_name = (data.get("full_name") or "").strip()
    role = data.get("role")
    technician_id = data.get("technician_id") or None

    errors = []
    if not username:
        errors.append("اسم المستخدم مطلوب")
    if not password or len(password) < 4:
        errors.append("كلمة المرور مطلوبة (4 أحرف على الأقل)")
    if not full_name:
        errors.append("الاسم الكامل مطلوب")
    if role not in ("admin", "technician"):
        errors.append("نوع الحساب غير صحيح")
    if role == "technician" and not technician_id:
        errors.append("يجب اختيار الفني المرتبط بهذا الحساب")
    if errors:
        return jsonify({"error": " - ".join(errors)}), 400

    db = get_db()
    if role == "technician":
        tech = db.execute("SELECT * FROM technicians WHERE id = ?", (technician_id,)).fetchone()
        if not tech:
            return jsonify({"error": "الفني المحدد غير موجود"}), 404
        linked = db.execute(
            "SELECT username FROM users WHERE technician_id = ? AND is_active = 1",
            (technician_id,),
        ).fetchone()
        if linked:
            return jsonify({"error": f"هذا الفني مرتبط مسبقاً بحساب '{linked['username']}'"}), 409

    try:
        cur = db.execute(
            "INSERT INTO users (username, password_hash, full_name, role, technician_id) "
            "VALUES (?, ?, ?, ?, ?)",
            (
                username, generate_password_hash(password), full_name, role,
                technician_id if role == "technician" else None,
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "اسم المستخدم مستخدم مسبقاً"}), 409

    row = db.execute(USER_SELECT + " WHERE u.id = ?", (cur.lastrowid,)).fetchone()
    return jsonify(serialize_user(row)), 201


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404

    data = request.get_json(force=True) or {}
    fields = {}

    if "full_name" in data:
        full_name = (data.get("full_name") or "").strip()
        if not full_name:
            return jsonify({"error": "الاسم الكامل مطلوب"}), 400
        fields["full_name"] = full_name

    if "username" in data:
        username = (data.get("username") or "").strip()
        if not username:
            return jsonify({"error": "اسم المستخدم مطلوب"}), 400
        fields["username"] = username

    new_role = data.get("role", user["role"])
    if new_role not in ("admin", "technician"):
        return jsonify({"error": "نوع الحساب غير صحيح"}), 400

    if (user["role"] == "admin" and new_role != "admin"
            and count_active_admins(db, exclude_id=user_id) == 0):
        return jsonify({"error": "لا يمكن تغيير دور آخر مدير نشط في النظام"}), 409

    if new_role == "technician":
        technician_id = data.get("technician_id", user["technician_id"])
        if not technician_id:
            return jsonify({"error": "يجب اختيار الفني المرتبط بهذا الحساب"}), 400
        tech = db.execute("SELECT * FROM technicians WHERE id = ?", (technician_id,)).fetchone()
        if not tech:
            return jsonify({"error": "الفني المحدد غير موجود"}), 404
        linked = db.execute(
            "SELECT username FROM users WHERE technician_id = ? AND is_active = 1 AND id != ?",
            (technician_id, user_id),
        ).fetchone()
        if linked:
            return jsonify({"error": f"هذا الفني مرتبط مسبقاً بحساب '{linked['username']}'"}), 409
        fields["technician_id"] = technician_id
    else:
        fields["technician_id"] = None

    fields["role"] = new_role

    if "is_active" in data:
        is_active = bool(data.get("is_active"))
        if not is_active and user["role"] == "admin" and count_active_admins(db, exclude_id=user_id) == 0:
            return jsonify({"error": "لا يمكن إلغاء تفعيل آخر مدير نشط في النظام"}), 409
        fields["is_active"] = int(is_active)

    if data.get("password"):
        if len(data["password"]) < 4:
            return jsonify({"error": "كلمة المرور قصيرة جداً (4 أحرف على الأقل)"}), 400
        fields["password_hash"] = generate_password_hash(data["password"])

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    try:
        db.execute(f"UPDATE users SET {set_clause} WHERE id = ?", list(fields.values()) + [user_id])
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "اسم المستخدم مستخدم مسبقاً"}), 409

    if user_id == session.get("user_id"):
        session["role"] = fields.get("role", session["role"])
        session["full_name"] = fields.get("full_name", session["full_name"])
        session["technician_id"] = fields.get("technician_id", session.get("technician_id"))

    row = db.execute(USER_SELECT + " WHERE u.id = ?", (user_id,)).fetchone()
    return jsonify(serialize_user(row))


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({"error": "المستخدم غير موجود"}), 404
    if user_id == session.get("user_id"):
        return jsonify({"error": "لا يمكنك حذف حسابك الخاص وأنت مسجل دخول به"}), 400
    if user["role"] == "admin" and user["is_active"] and count_active_admins(db, exclude_id=user_id) == 0:
        return jsonify({"error": "لا يمكن حذف آخر مدير نشط في النظام"}), 409
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# أدوات مساعدة لتحويل صف المهمة إلى قاموس كامل
# --------------------------------------------------------------------------
def serialize_task(db, row):
    branches = db.execute(
        """
        SELECT tb.id, tb.branch_id, tb.is_completed, tb.completed_at, b.name AS branch_name
        FROM task_branches tb JOIN branches b ON b.id = tb.branch_id
        WHERE tb.task_id = ?
        ORDER BY b.sort_order, b.id
        """,
        (row["id"],),
    ).fetchall()

    total = len(branches)
    done = sum(1 for b in branches if b["is_completed"])

    return {
        "id": row["id"],
        "task_code": row["task_code"],
        "title": row["title"],
        "task_type": row["task_type"],
        "maintenance_kind": row["maintenance_kind"],
        "zone_id": row["zone_id"],
        "zone_name": row["zone_name"],
        "technician_id": row["technician_id"],
        "technician_name": row["technician_name"],
        "start_date": row["start_date"],
        "status": row["status"],
        "target_all_branches": bool(row["target_all_branches"]),
        "notes": row["notes"],
        "created_at": row["created_at"],
        "completed_at": row["completed_at"],
        "branches": [
            {
                "id": b["branch_id"],
                "name": b["branch_name"],
                "is_completed": bool(b["is_completed"]),
                "completed_at": b["completed_at"],
            }
            for b in branches
        ],
        "progress": {"done": done, "total": total},
    }


TASK_SELECT = """
    SELECT t.*, z.name AS zone_name, tc.name AS technician_name
    FROM tasks t
    JOIN zones z ON z.id = t.zone_id
    LEFT JOIN technicians tc ON tc.id = t.technician_id
"""


# --------------------------------------------------------------------------
# المهام: قراءة / إنشاء / تعديل / حذف
# --------------------------------------------------------------------------
@app.route("/api/tasks", methods=["GET"])
@login_required
def list_tasks():
    db = get_db()
    filters = []
    params = []

    status = request.args.get("status")
    if status and status in STATUS_LIST:
        filters.append("t.status = ?")
        params.append(status)

    task_type = request.args.get("task_type")
    if task_type in TASK_TYPES:
        filters.append("t.task_type = ?")
        params.append(task_type)

    zone_id = request.args.get("zone_id")
    if zone_id:
        filters.append("t.zone_id = ?")
        params.append(zone_id)

    technician_id = request.args.get("technician_id")
    if technician_id:
        filters.append("t.technician_id = ?")
        params.append(technician_id)

    maintenance_kind = request.args.get("maintenance_kind")
    if maintenance_kind in MAINTENANCE_KINDS:
        filters.append("t.maintenance_kind = ?")
        params.append(maintenance_kind)

    search = request.args.get("q")
    if search:
        filters.append("(t.title LIKE ? OR t.task_code LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])

    # الفني يرى مهامه فقط
    if session.get("role") == "technician" and session.get("technician_id"):
        filters.append("t.technician_id = ?")
        params.append(session["technician_id"])

    query = TASK_SELECT
    if filters:
        query += " WHERE " + " AND ".join(filters)
    query += " ORDER BY t.created_at DESC"

    rows = db.execute(query, params).fetchall()
    tasks = [serialize_task(db, r) for r in rows]

    return jsonify(tasks)


@app.route("/api/tasks/<int:task_id>", methods=["GET"])
@login_required
def get_task(task_id):
    db = get_db()
    row = db.execute(TASK_SELECT + " WHERE t.id = ?", (task_id,)).fetchone()
    if not row:
        return jsonify({"error": "المهمة غير موجودة"}), 404
    return jsonify(serialize_task(db, row))


@app.route("/api/tasks", methods=["POST"])
@admin_required
def create_task():
    data = request.get_json(force=True) or {}

    title = (data.get("title") or "").strip()
    task_type = data.get("task_type")
    maintenance_kind = data.get("maintenance_kind")
    zone_id = data.get("zone_id")
    technician_id = data.get("technician_id")
    start_date = data.get("start_date")
    notes = (data.get("notes") or "").strip()
    target_all_branches = bool(data.get("target_all_branches"))
    branch_ids = data.get("branch_ids") or []

    errors = []
    if task_type not in TASK_TYPES:
        errors.append("نوع المهمة غير صحيح")
    if maintenance_kind not in MAINTENANCE_KINDS:
        errors.append("نوع الصيانة غير صحيح")
    if not zone_id:
        errors.append("يجب تحديد المنطقة المستهدفة")
    if not start_date:
        errors.append("يجب تحديد تاريخ البدء")
    if not target_all_branches and not branch_ids:
        errors.append("يجب تحديد فرع واحد على الأقل أو اختيار كل الفروع")

    if errors:
        return jsonify({"error": " - ".join(errors)}), 400

    db = get_db()
    zone = db.execute("SELECT * FROM zones WHERE id = ?", (zone_id,)).fetchone()
    if not zone:
        return jsonify({"error": "المنطقة المستهدفة غير موجودة"}), 404

    if not title:
        title = f"صيانة وقائية للكاميرات - {zone['name']}" if task_type == "camera" else f"مهمة {maintenance_kind} - {zone['name']}"

    task_code = gen_task_code(db, task_type)

    cur = db.execute(
        """
        INSERT INTO tasks (task_code, title, task_type, maintenance_kind, zone_id,
                            technician_id, start_date, status, target_all_branches,
                            notes, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'قيد الانتظار', ?, ?, ?)
        """,
        (
            task_code, title, task_type, maintenance_kind, zone_id,
            technician_id or None, start_date, int(target_all_branches),
            notes, session.get("user_id"),
        ),
    )
    task_id = cur.lastrowid

    if target_all_branches:
        all_branches = db.execute(
            "SELECT id FROM branches WHERE zone_id = ?", (zone_id,)
        ).fetchall()
        branch_ids = [b["id"] for b in all_branches]

    for bid in branch_ids:
        db.execute(
            "INSERT OR IGNORE INTO task_branches (task_id, branch_id) VALUES (?, ?)",
            (task_id, bid),
        )

    db.commit()

    row = db.execute(TASK_SELECT + " WHERE t.id = ?", (task_id,)).fetchone()
    return jsonify(serialize_task(db, row)), 201


@app.route("/api/tasks/<int:task_id>", methods=["PUT"])
@admin_required
def update_task(task_id):
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return jsonify({"error": "المهمة غير موجودة"}), 404

    data = request.get_json(force=True) or {}
    fields = {}

    for key in ["title", "maintenance_kind", "technician_id", "start_date", "notes"]:
        if key in data:
            fields[key] = data[key]

    if "maintenance_kind" in fields and fields["maintenance_kind"] not in MAINTENANCE_KINDS:
        return jsonify({"error": "نوع الصيانة غير صحيح"}), 400

    if fields:
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        db.execute(
            f"UPDATE tasks SET {set_clause} WHERE id = ?",
            list(fields.values()) + [task_id],
        )
        db.commit()

    row = db.execute(TASK_SELECT + " WHERE t.id = ?", (task_id,)).fetchone()
    return jsonify(serialize_task(db, row))


@app.route("/api/tasks/<int:task_id>", methods=["DELETE"])
@admin_required
def delete_task(task_id):
    db = get_db()
    task = db.execute("SELECT id FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return jsonify({"error": "المهمة غير موجودة"}), 404
    db.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    db.commit()
    return jsonify({"ok": True})


# --------------------------------------------------------------------------
# تغيير حالة المهمة (مع نافذة تحديد الفروع المنجزة عند الإكمال)
# --------------------------------------------------------------------------
@app.route("/api/tasks/<int:task_id>/status", methods=["PUT"])
@login_required
def update_status(task_id):
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return jsonify({"error": "المهمة غير موجودة"}), 404

    if session.get("role") == "technician" and task["technician_id"] != session.get("technician_id"):
        return jsonify({"error": "لا يمكنك تعديل مهمة غير مخصصة لك"}), 403

    data = request.get_json(force=True) or {}
    new_status = data.get("status")
    if new_status not in STATUS_LIST:
        return jsonify({"error": "حالة غير صحيحة"}), 400

    if new_status == "مكتمل":
        completed_branch_ids = data.get("completed_branch_ids")
        complete_all = bool(data.get("complete_all_branches"))

        all_task_branches = db.execute(
            "SELECT branch_id FROM task_branches WHERE task_id = ?", (task_id,)
        ).fetchall()
        all_ids = [b["branch_id"] for b in all_task_branches]

        if complete_all:
            completed_branch_ids = all_ids
        elif completed_branch_ids is None:
            return jsonify({"error": "يجب تحديد الفروع التي تم إنجازها"}), 400

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # إعادة ضبط جميع الفروع كغير منجزة أولاً
        db.execute(
            "UPDATE task_branches SET is_completed = 0, completed_at = NULL WHERE task_id = ?",
            (task_id,),
        )
        for bid in completed_branch_ids:
            db.execute(
                "UPDATE task_branches SET is_completed = 1, completed_at = ? "
                "WHERE task_id = ? AND branch_id = ?",
                (now, task_id, bid),
            )

        db.execute(
            "UPDATE tasks SET status = ?, completed_at = ? WHERE id = ?",
            (new_status, now, task_id),
        )
    else:
        db.execute(
            "UPDATE tasks SET status = ?, completed_at = NULL WHERE id = ?",
            (new_status, task_id),
        )

    db.commit()
    row = db.execute(TASK_SELECT + " WHERE t.id = ?", (task_id,)).fetchone()
    return jsonify(serialize_task(db, row))


# --------------------------------------------------------------------------
# لوحة الإحصائيات السريعة
# --------------------------------------------------------------------------
@app.route("/api/stats", methods=["GET"])
@login_required
def stats():
    db = get_db()
    base = "SELECT COUNT(*) AS c FROM tasks"
    where = ""
    params = []
    if session.get("role") == "technician" and session.get("technician_id"):
        where = " WHERE technician_id = ?"
        params = [session["technician_id"]]

    total = db.execute(base + where, params).fetchone()["c"]

    def count_status(s):
        q = base + (" WHERE status = ?" if not where else where + " AND status = ?")
        return db.execute(q, params + [s]).fetchone()["c"]

    def count_type(t):
        q = base + (" WHERE task_type = ?" if not where else where + " AND task_type = ?")
        return db.execute(q, params + [t]).fetchone()["c"]

    return jsonify({
        "total": total,
        "pending": count_status("قيد الانتظار"),
        "in_progress": count_status("قيد التنفيذ"),
        "completed": count_status("مكتمل"),
        "camera_tasks": count_type("camera"),
        "normal_tasks": count_type("normal"),
    })


# --------------------------------------------------------------------------
# التصدير: Excel و PDF
# --------------------------------------------------------------------------
def fetch_export_rows(db, only_camera=False, status=None, zone_id=None,
                       date_from=None, date_to=None):
    filters = []
    params = []
    if only_camera:
        filters.append("t.task_type = 'camera'")
    if status:
        filters.append("t.status = ?")
        params.append(status)
    if zone_id:
        filters.append("t.zone_id = ?")
        params.append(zone_id)
    if date_from:
        filters.append("t.start_date >= ?")
        params.append(date_from)
    if date_to:
        filters.append("t.start_date <= ?")
        params.append(date_to)

    query = TASK_SELECT
    if filters:
        query += " WHERE " + " AND ".join(filters)
    query += " ORDER BY t.start_date DESC"

    rows = db.execute(query, params).fetchall()
    return [serialize_task(db, r) for r in rows]


@app.route("/api/export/excel", methods=["GET"])
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    only_camera = request.args.get("type") == "camera"
    status = request.args.get("status") or None
    zone_id = request.args.get("zone_id") or None

    db = get_db()
    tasks = fetch_export_rows(db, only_camera=only_camera, status=status, zone_id=zone_id)

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "الصيانة الوقائية للكاميرات" if only_camera else "المهام"

    headers = [
        "كود المهمة", "العنوان", "النوع", "نوع الصيانة", "المنطقة",
        "الفني", "تاريخ البدء", "الحالة", "إجمالي الفروع",
        "الفروع المنجزة", "تاريخ الإكمال", "الفروع المستهدفة",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    type_labels = {"camera": "صيانة وقائية كاميرات", "normal": "مهمة عادية"}

    for t in tasks:
        branch_names = "، ".join(b["name"] for b in t["branches"])
        ws.append([
            t["task_code"], t["title"], type_labels.get(t["task_type"], t["task_type"]),
            t["maintenance_kind"], t["zone_name"], t["technician_name"] or "-",
            t["start_date"], t["status"], t["progress"]["total"],
            t["progress"]["done"], t["completed_at"] or "-", branch_names,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [16, 30, 18, 12, 16, 16, 13, 13, 12, 12, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "تقرير_الصيانة_الوقائية_للكاميرات.xlsx" if only_camera else "تقرير_المهام.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/export/pdf", methods=["GET"])
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display

    only_camera = request.args.get("type") == "camera"
    status = request.args.get("status") or None
    zone_id = request.args.get("zone_id") or None

    db = get_db()
    tasks = fetch_export_rows(db, only_camera=only_camera, status=status, zone_id=zone_id)

    font_path = os.path.join(os.path.dirname(__file__), "static", "fonts", "NotoNaskhArabic-Regular.ttf")
    font_bold_path = os.path.join(os.path.dirname(__file__), "static", "fonts", "NotoNaskhArabic-Bold.ttf")
    pdfmetrics.registerFont(TTFont("Arabic", font_path))
    pdfmetrics.registerFont(TTFont("Arabic-Bold", font_bold_path if os.path.exists(font_bold_path) else font_path))

    def ar(text):
        if text is None:
            return ""
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
    )

    title_style = ParagraphStyle(
        "title", fontName="Arabic-Bold", fontSize=16, alignment=1, textColor=colors.HexColor("#1E293B"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "sub", fontName="Arabic", fontSize=10, alignment=1, textColor=colors.HexColor("#64748B"),
    )

    elements = []
    title_text = "تقرير الصيانة الوقائية للكاميرات" if only_camera else "تقرير المهام"
    elements.append(Paragraph(ar(title_text), title_style))
    elements.append(Paragraph(ar(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), sub_style))
    elements.append(Spacer(1, 0.5 * cm))

    type_labels = {"camera": "صيانة كاميرات", "normal": "مهمة عادية"}

    headers = ["الفروع المنجزة/الكل", "تاريخ البدء", "الفني", "المنطقة",
               "نوع الصيانة", "النوع", "العنوان", "الكود"]
    data = [[ar(h) for h in headers]]

    for t in tasks:
        prog = f"{t['progress']['done']}/{t['progress']['total']}"
        data.append([
            ar(prog),
            ar(t["start_date"]),
            ar(t["technician_name"] or "-"),
            ar(t["zone_name"]),
            ar(t["maintenance_kind"]),
            ar(type_labels.get(t["task_type"], t["task_type"])),
            ar(t["title"]),
            ar(t["task_code"]),
        ])

    table = Table(data, repeatRows=1, colWidths=[2.6*cm, 2.6*cm, 3.0*cm, 3.0*cm, 2.6*cm, 3.2*cm, 6.5*cm, 3.4*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Arabic-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Arabic"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
    buf.seek(0)

    filename = "تقرير_الصيانة_الوقائية_للكاميرات.pdf" if only_camera else "تقرير_المهام.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")


# --------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
